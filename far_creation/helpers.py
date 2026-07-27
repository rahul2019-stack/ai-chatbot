import csv
import pandas as pd
from config import config_dict
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    Font
)
from openpyxl.utils import get_column_letter

def create_excel_sheet(env, app, department, bastion_node_ips, bootstrap_node_ips, master_node_ips, worker_node_ips, infra_node_ips, vip1, vip2, dns_name, vcentre_name):
    excel_file_name = f"far_{department}_{env}_{app}.xlsx"
    description_df = _form_description_data_df(env, app, department, bastion_node_ips, bootstrap_node_ips, master_node_ips, worker_node_ips, infra_node_ips, vip1, vip2)
    lb_df = _form_lb_data_df(env, app, department, bootstrap_node_ips, master_node_ips, infra_node_ips, dns_name)
    dns_entries_df = _form_dns_entries_df(env, app, department, bootstrap_node_ips, master_node_ips, infra_node_ips, worker_node_ips, bastion_node_ips, vip1, vip2, dns_name)
    far_entries_df = _form_far_entries_df(env, app, department, bastion_node_ips, bootstrap_node_ips, master_node_ips, worker_node_ips, infra_node_ips, vip1, vip2, dns_name, vcentre_name)
    
    with pd.ExcelWriter(excel_file_name) as writer:
        description_df.to_excel(writer, sheet_name="Description", index=False)
        lb_df.to_excel(writer, sheet_name="LB_Details", index=False)
        dns_entries_df.to_excel(writer, sheet_name="DNS_Entry", index=False)
        far_entries_df.to_excel(writer, sheet_name="FAR_Rules", index=False)

    _format_sheet(
        excel_file=excel_file_name,
        sheet_name="LB_Details",
        merge_columns=["A"]
    )
    _format_sheet(
        excel_file=excel_file_name,
        sheet_name="Description"
    )
    _format_sheet(
        excel_file=excel_file_name,
        sheet_name="DNS_Entry"
    )
    _format_sheet(
        excel_file=excel_file_name,
        sheet_name="FAR_Rules"
    )

def _form_far_entries_df(env, app, department, bastion_node_ips, bootstrap_node_ips, master_node_ips, worker_node_ips, infra_node_ips, vip1, vip2, dns_name, vcentre_name):
    serial_number = list(range(1, 21))

    if env.lower() == "uat" or env.lower() == "pre-prod" or env.lower() == "dev":
        central_bastion_ip = config_dict["central_meghdoot_bastion_node_ip_uat"]
        central_mirror_ip = config_dict["central_meghdoot_mirror_node_ip_uat"]
    else:
        central_bastion_ip = config_dict["central_meghdoot_bastion_node_ip_prod"]
        central_mirror_ip = config_dict["central_meghdoot_mirror_node_ip_prod"]
    meghdoot_central_nodes = f"{central_bastion_ip}\n{central_mirror_ip}"
    
    source_ip_rule_1_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(bastion_node_ips),
        "\n".join(master_node_ips),
        "\n".join(infra_node_ips),
        "\n".join(worker_node_ips),
        meghdoot_central_nodes
    ]
    source_ip_rule_1 = "\n".join(source_ip_rule_1_list)
    destination_ip_rule_1 = "\n".join(source_ip_rule_1_list)
    service_ports1 = [
        "TCP/1936",
        "TCP/9000-9999",
        "TCP/10249-10259",
        "TCP/10256",
        "TCP/30000-32767",
        "UDP/30000-32767",
        "UDP/30000-32767",
        "UDP/4789",
        "UDP/4500",
        "UDP/500",
        "UDP/6081",
        "UDP/9000-9999",
    ]
    service_rule_1 = "\n".join(service_ports1)

    source_ip_rule_2 = "\n".join(source_ip_rule_1_list)
    destination_ip_rule_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(master_node_ips)
    ]
    destination_ip_rule_2 = "\n".join(destination_ip_rule_list)
    service_rule_2 = "TCP/6443"

    source_ip_rule_3_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(master_node_ips)
    ]
    source_ip_rule_3 = "\n".join(source_ip_rule_3_list)
    destination_ip_rule_3 = "\n".join(source_ip_rule_3_list)
    service_rule_3 = "TCP/2379-2380"

    source_ip_rule_4_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(master_node_ips),
        "\n".join(infra_node_ips),
        "\n".join(worker_node_ips),
        "\n".join(bastion_node_ips),
        central_bastion_ip
    ]
    source_ip_rule_4 = "\n".join(source_ip_rule_4_list)
    destination_ip_rule_4 =f"{central_mirror_ip}"
    service_rule_4 = "TCP/8443"

    source_ip_rule_5_list = [
        "\n".join(bastion_node_ips),
        "\n".join(bootstrap_node_ips),
        "\n".join(master_node_ips),
        "\n".join(infra_node_ips),
        "\n".join(worker_node_ips),
        central_mirror_ip
    ]
    source_ip_rule_5 = "\n".join(source_ip_rule_5_list)
    destination_ip_rule_5_list = [
        central_bastion_ip,
        "\n".join(bastion_node_ips)
    ]
    destination_ip_rule_5 = "\n".join(destination_ip_rule_5_list)
    service_rule_5 = "TCP/9090 \nTCP/8080"

    source_ip_rule_6 = "\n".join(source_ip_rule_1_list)
    destination_ip_rule_6 = vip1
    service_rule_6 = "TCP/6443 \nTCP/22623"

    source_ip_rule_7 = "\n".join(source_ip_rule_1_list)
    destination_ip_rule_7 = vip2
    service_rule_7 = "TCP/443"

    # To allow comm between VIP1 API server to bootstrap and master nodes
    source_ip_rule_8 = vip1
    destination_ip_rule_8_list = [
        "\n".join(bootstrap_node_ips)
        "\n".join(master_node_ips)
    ]
    destination_ip_rule_8 = "\n".join(destination_ip_rule_8_list)
    service_rule_8 = "TCP/6443 \nTCP/22623"

    # To allow comm between *apps VIP2 to infra nodes, if no infra nodes are there traffic should go to master nodes
    source_ip_rule_9 = vip2
    destination_ip_rule_9_list = []
    if len(infra_node_ips) == 0:
        destination_ip_rule_9_list.append("\n".join(bootstrap_node_ips))
        destination_ip_rule_9_list.append("\n".join(master_node_ips))
    else:
        destination_ip_rule_9_list.append("\n".join(infra_node_ips))
    destination_ip_rule_9 = "\n".join(destination_ip_rule_9_list)
    service_rule_9 = "TCP/443"

    source_ip_rule_10_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(bastion_node_ips),
        "\n".join(master_node_ips),
        "\n".join(infra_node_ips),
        "\n".join(worker_node_ips),
        meghdoot_central_nodes
    ]
    source_ip_rule_10 = "\n".join(source_ip_rule_10_list)
    vcentre_ips = _get_vcentre_ips(vcentre_name)
    destination_ip_rule_10 = "\n".join(vcentre_ips)
    service_rule_10 = "TCP/443"

    source_ip_rule_11_list = [
        "\n".join(bastion_node_ips),
        meghdoot_central_nodes
    ]
    source_ip_rule_11 = "\n".join(source_ip_rule_11_list)
    destination_ip_rule_11_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(bastion_node_ips),
        "\n".join(master_node_ips),
        "\n".join(infra_node_ips),
        "\n".join(worker_node_ips),
        meghdoot_central_nodes
    ]
    destination_ip_rule_11 = "\n".join(destination_ip_rule_11_list)
    service_rule_11 = "TCP/22"

    source_ip_rule_12_list = [
        "\n".join(bootstrap_node_ips),
        "\n".join(bastion_node_ips),
        "\n".join(master_node_ips),
        "\n".join(infra_node_ips),
        "\n".join(worker_node_ips),
        meghdoot_central_nodes
    ]
    source_ip_rule_12 = "\n".join(source_ip_rule_12_list)
    destination_ip_rule_12 = "\n".join(config_dict["NTP_servers"])
    service_rule_12 = "UDP/123"

    source_ip_rule_13 = source_ip_rule_12
    destination_ip_rule_13 = "SMTP IP or URL"
    service_rule_13 = "SMTP Port"

    source_ip_rule_14 = source_ip_rule_12
    destination_ip_rule_14 = "Syslog IP or URL"
    service_rule_14 = "Syslog Port"

    source_ip_rule_15 = source_ip_rule_12
    destination_ip_rule_15 = "NFS IP or URL"
    service_rule_15 = "NFS Port"

    source_ip_rule_16 = "\n".join(master_node_ips)
    destination_ip_rule_16 = "\n".join(config_dict.get("AD_auth_ips"))
    service_rule_16 = "TCP/636"

    source_ip_rule_17 = "\n".join(config_dict.get("ocp_team_desktop_ip"))
    destination_ip_rule_17 = "\n".join(bastion_node_ips) + "\n" + central_bastion_ip
    service_rule_17 = "TCP/22"

    source_ip_rule_18 = source_ip_rule_17
    destination_ip_rule_18 = central_mirror_ip
    service_rule_18 = "TCP/8443"

    source_ip_rule_19 = source_ip_rule_17
    destination_ip_rule_19 = vip2
    service_rule_19 = "TCP/443"

    source_ip_rule_20 = source_ip_rule_12
    destination_ip_rule_20 = "S3 bucket URL"
    service_rule_20 = "TCP/443"




    data = {
        "Sr.No": list(range(21)),
        "Source IP Address": [source_ip_rule_1, source_ip_rule_2, source_ip_rule_3, source_ip_rule_4, source_ip_rule_5,
                              source_ip_rule_6, source_ip_rule_7, source_ip_rule_8, source_ip_rule_9, source_ip_rule_10,
                              source_ip_rule_11, source_ip_rule_12, source_ip_rule_13, source_ip_rule_14, source_ip_rule_15,
                              source_ip_rule_16, source_ip_rule_17, source_ip_rule_18, source_ip_rule_19, source_ip_rule_20],
        "User": ["Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any",
                 "Any", "Any", "Any", "Any"],
        "Destination IP Address": [destination_ip_rule_1, destination_ip_rule_2, destination_ip_rule_3, destination_ip_rule_4, destination_ip_rule_5,
                                   destination_ip_rule_6, destination_ip_rule_7, destination_ip_rule_8, destination_ip_rule_9,
                                   destination_ip_rule_10, destination_ip_rule_11, destination_ip_rule_12, destination_ip_rule_13,
                                   destination_ip_rule_14, destination_ip_rule_15, destination_ip_rule_16,
                                   destination_ip_rule_17, destination_ip_rule_18, destination_ip_rule_19, destination_ip_rule_20],
        "Application": ["Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any", 
                        "Any", "Any", "Any", "Any", "Any", "Any", "Any", "Any"],
        "Service": [service_rule_1, service_rule_2, service_rule_3, service_rule_4, service_rule_5, service_rule_6, service_rule_7,
                    service_rule_8, service_rule_9, service_rule_10, service_rule_11, service_rule_12, service_rule_13, service_rule_14,
                    service_rule_15, service_rule_16, service_rule_17, service_rule_18, service_rule_19, service_rule_20],
        "Action": ["Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "Allow",
                   "Allow", "Allow", "Allow", "Allow", "Allow", "Allow", "ALlow"],
        "Comment": ["This rule allows all traffic between the OpenShift nodes and the central bastion/mirror nodes.", 
                    "This rule allows traffic from the OpenShift nodes to the API server on the bootstrap and master nodes.", 
                    "This rule allows ETCD sync traffic between the bootstrap and master nodes.",
                    "Allow all nodes to pull container images from central mirror",
                    "Bastion node hosts the ignition config files and this rule is needed at the time of cluster installation to download ignition configs while creating RHCOS VMs",
                    "To connect API and API-Int url from all OCP nodes",
                    "To connect *.apps from all OCP nodes",
                    "To allow communication from cluster nodes to respective vcentre/vsphere",
                    "To allow SSH login to all cluster nodes from bastion nodes",
                    "For clock sync among all OCP nodes with NTP servers",
                    "Send openshift alerts from all OCP nodes to mailbox",
                    "To allow all communication from all nodes to all cluster nodes to forward logs",
                    "To allow all nodes to access NFS storage",
                    "Openshift platform authentication from AD",
                    "SSH login to bastion nodes",
                    "SSH login to miiror node",
                    "To access web console of app from OCP Admin team's desktop",
                    "For loki and quay setup"],
        "Remarks": ["", "", "", "", "", "", "",  "", "", "", "", "", "", "", "", "",  "", "", "", "" ]
    }
    print(f"FAR rule data formed is {data}")
    return pd.DataFrame(data)

def _get_vcentre_ips(vcentre_name):
    if "b1u" in vcentre_name.lower():
        vcentre_ips = config_dict["vcentre_ips_b1u"]
    elif "b1p" in vcentre_name.lower():
        vcentre_ips = config_dict["vcentre_ips_b1p"]
    elif "a2p" in vcentre_name.lower():
        vcentre_ips = config_dict["vcentre_ips_a2p"]
    else:
        raise Exception("The vcentre IP's are not configured, Kindly change in config.py file")
    return vcentre_ips

def _form_dns_entries_df(env, app, department, bootstrap_node_ips, master_node_ips, infra_node_ips, worker_node_ips, bastion_node_ips, vip1, vip2, dns_name):
    total_vms = len(bootstrap_node_ips) + len(master_node_ips) + len(infra_node_ips) + len(worker_node_ips) + len(bastion_node_ips)
    serials = list(range(1, total_vms+4))  # +4 for VIP1, VIP1 ,VIP2, and one extra to account for the starting index of 1
    Vm_names = ["VIP1", "VIP1", "VIP2"]
    Vm_names.extend(["AO to fillup"] * total_vms)
    ip_address = [vip1, vip1, vip2]
    for ip in bootstrap_node_ips:
        ip_address.append(ip)
    for ip in bastion_node_ips:
        ip_address.append(ip)
    for ip in master_node_ips:
        ip_address.append(ip)
    for ip in infra_node_ips:
        ip_address.append(ip)
    for ip in worker_node_ips:
        ip_address.append(ip)
    backend_dns = [f"api.{dns_name}.bank.sbi", f"api.{dns_name}.bank.sbi", f"*.apps.{dns_name}.bank.sbi"]
    bootstrap_dns = f"bootstrap.{dns_name}.bank.sbi"
    backend_dns.extend([bootstrap_dns])
    for i in range(len(bastion_node_ips)):
        if len(bastion_node_ips) == 1:
            backend_dns.append(f"bastion.{dns_name}.bank.sbi")
        else:
            backend_dns.append(f"bastion{i+1}.{dns_name}.bank.sbi")
    for i in range(len(master_node_ips)):
        backend_dns.append(f"master{i+1}.{dns_name}.bank.sbi")
    for i in range(len(infra_node_ips)):
        backend_dns.append(f"infra{i+1}.{dns_name}.bank.sbi")
    for i in range(len(worker_node_ips)):
        backend_dns.append(f"worker{i+1}.{dns_name}.bank.sbi")

    data = {
        "Sr.No": serials,
        "VM Name": Vm_names,
        "DNS Entry": backend_dns,
        "IP Address": ip_address
    }

    print(f"Data formed for DNS entries is {data}")

    return pd.DataFrame(data)

def _format_sheet(
    excel_file,
    sheet_name,
    merge_columns=None
):

    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True    
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    if merge_columns:

        for column in merge_columns:

            start_row = 2

            while start_row <= ws.max_row:

                value = ws[f"{column}{start_row}"].value
                end_row = start_row

                while (
                    end_row < ws.max_row
                    and ws[f"{column}{end_row+1}"].value == value
                ):
                    end_row += 1

                if end_row > start_row:
                    ws.merge_cells(
                        f"{column}{start_row}:{column}{end_row}"
                    )

                    ws[f"{column}{start_row}"].alignment = center_alignment

                start_row = end_row + 1

    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 3

    wb.save(excel_file)

def _form_lb_data_df(env, app, department, bootstrap_node_ips, master_node_ips, infra_node_ips, dns_name):
    total_bootstrap_and_master_nodes = len(bootstrap_node_ips) + len(master_node_ips)
    total_infra_nodes = len(infra_node_ips)
    total_master_nodes = len(master_node_ips)
    api_server = f"api.{dns_name}.bank.sbi"
    api_int_server = f"api-int.{dns_name}.bank.sbi"
    apps_server = f"*.apps.{dns_name}.bank.sbi"

    front_back_end_port = []
    front_back_end_port.extend(['6443'] * total_bootstrap_and_master_nodes)
    front_back_end_port.extend(['22623'] * total_bootstrap_and_master_nodes)
    front_back_end_port.extend(['443'] * total_infra_nodes)

    front_end_list = []
    front_end_list.extend([api_server] * total_bootstrap_and_master_nodes)
    front_end_list.extend([api_int_server] * total_bootstrap_and_master_nodes)
    front_end_list.extend([apps_server] * total_infra_nodes)

    # backend_dns = [bootstrap_dns, master_dns * total_master_nodes, bootstrap_dns, master_dns * total_master_nodes, infra_dns * total_infra_nodes]
    backend_dns = []
    bootstrap_dns = f"bootstrap.{dns_name}.bank.sbi"
    backend_dns.extend([bootstrap_dns])
    for i in range(total_master_nodes):
        backend_dns.append(f"master{i+1}.{dns_name}.bank.sbi")
    backend_dns.extend([bootstrap_dns])
    for i in range(total_master_nodes):
        backend_dns.append(f"master{i+1}.{dns_name}.bank.sbi")
    for i in range(total_infra_nodes):
        backend_dns.append(f"infra{i+1}.{dns_name}.bank.sbi")

    data = {
        "FrontEnd": front_end_list,
        "FrontEndPort": front_back_end_port,
        "Backend": backend_dns,
        "BackendPort": front_back_end_port,
    }

    print(f"Data formed for LB details is {data}")
    df = pd.DataFrame(data)
    return df

def _form_description_data_df(env, app, department, bastion_node_ips, bootstrap_node_ips, master_node_ips, worker_node_ips, infra_node_ips, vip1, vip2):
    node_ips = ""
    for i, node_ip in enumerate(bastion_node_ips, start=1):
        if len(bastion_node_ips) == 1:
            node_ips += f" Bastion/Registry {node_ip}\n"
        else:
            node_ips = f" Registry {i} {node_ip}\n"
    for i, node_ip in enumerate(bootstrap_node_ips, start=1):
        if len(bootstrap_node_ips) == 1:
            node_ips += f" Bootstrap {node_ip}\n"
        else:
            node_ips += f" Bootstrap{i} {node_ip}\n"
    for i, node_ip in enumerate(master_node_ips, start=1):
        if len(master_node_ips) == 1:
            node_ips += f" Master {node_ip}\n"
        else:
            node_ips += f" Master{i} {node_ip}\n"
    for i, node_ip in enumerate(worker_node_ips, start=1):
        if len(worker_node_ips) == 1:
            node_ips += f" Worker {node_ip}\n"
        else:
            node_ips += f" Worker{i} {node_ip}\n"
    if len(infra_node_ips) > 0:
        for i, node_ip in enumerate(infra_node_ips, start=1):
            if len(infra_node_ips) == 1:
                node_ips += f" Infra {node_ip}\n"
            else:
                node_ips += f" Infra{i} {node_ip}\n"

    if env.lower() == "uat" or env.lower() == "pre-prod" or env.lower() == "dev":
        node_ips += f" Central Bastion {config_dict['central_meghdoot_bastion_node_ip_uat']}\n"
        node_ips += f" Central Mirror {config_dict['central_meghdoot_mirror_node_ip_uat']}\n"
    else:
        node_ips += f" Central Bastion {config_dict['central_meghdoot_bastion_node_ip_prod']}\n"
        node_ips += f" Central Mirror {config_dict['central_meghdoot_mirror_node_ip_prod']}\n"

    vip_str = f"VIP1: {vip1}\nVIP2: {vip2}"
    ldap_ip_str = "\n".join(config_dict["LDAP_ip"])
    ntp_ip_str = "\n".join(config_dict["NTP_ip"])
    ocp_team_desktop_ip_str = "\n".join(config_dict["ocp_team_desktop_ip"])

    data = {
        "Host-Description": ["Node Info", "LB IP", "LDAP url", "LDAP IP", "NTP IP", "Windows IP"],
        "Host-Ips": [node_ips, vip_str, config_dict["LDAP_url"], ldap_ip_str, ntp_ip_str, ocp_team_desktop_ip_str],
        "Details": ["Node Ip info", "VIP info", "This url is used to establish connection between Ocp cluster and vcenter platform as well as authentication",
                    "Used for authentication", "Used for time sync", "Meghdoot OCP Team's desktop IP"]
    }
    df = pd.DataFrame(data)
    return df